import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from netaddr import cidr_merge

sys.path.append("/home/ngsop/lilaApp/plugins/utilidadesPlugins")
from loggingConfig import LoggerFileConfig
from constantesPlugins import LOG_CONFIG_FILES
logging = LoggerFileConfig().crearLogFile(LOG_CONFIG_FILES.get("blacklist_check"))

from constants import constantes

def juntar_sub_bloques(lista_ips):

    return [str(res) for res in cidr_merge(lista_ips)]

def generar_reporte(reporte_general):
    try:
        logging.info(f"Generando reporte")
        nombre_archivo = ("prueba_blacklist.xlsx")
        wb = Workbook()
        negrita = Font(bold=True)

        # Recolectar dominios
        todos_los_dominios = set()
        for contenido in reporte_general.values():
            for datos in contenido.get("bloques", {}).values():
                if datos["resultado"] == "AUDITORIA":
                    for h in datos.get("ips", []) or []:
                        if h and isinstance(h.get("dominios"), str):
                            todos_los_dominios.update(h["dominios"].split(", "))
        columnas_dnsbl = sorted(todos_los_dominios)

        # Agregar columnas
        constantes.PESTANAS["AUDITORIA"]["columnas"] = ["Bloque", "IP", "Resultado"] + columnas_dnsbl

        #Generar pestañas
        hojas = {}
        primera = True
        for nombre, config in constantes.PESTANAS.items():
            ws = wb.active if primera else wb.create_sheet(nombre)
            if primera:
                ws.title = nombre
                primera = False

            hojas[nombre] = ws
            color = PatternFill("solid", fgColor=config["color"])

            ws.append(config["columnas"])
            for celda in ws[1]:
                celda.font = negrita

            for red_original, contenido in reporte_general.items():
                for segmento, datos in contenido.get("bloques", {}).items():
                    if datos["resultado"] != nombre:
                        continue

                    if nombre == "AUDITORIA":
                        if not datos.get("ips"):
                            fila = [segmento, "N/A", nombre] + [""] * len(columnas_dnsbl)
                            ws.append(fila)
                            for celda in ws[ws.max_row]:
                                celda.fill = color
                            continue

                        for h in datos["ips"]:
                            if not h:
                                continue
                            ws.append([segmento, h["ip"], nombre])

                            for celda in ws[ws.max_row]:
                                celda.fill = color
                    else:
                        ws.append([segmento, nombre])
                        for celda in ws[ws.max_row]:
                            celda.fill = color

        #Pestaña RESUMEN
        generar_resumen(wb, reporte_general)

        wb.save(nombre_archivo)

        logging.info(f"Reporte generado exitosamente: {nombre_archivo}")

        return nombre_archivo
    
    except Exception as e:
        print("ERROR:", str(e))

def generar_resumen(wb, reporte_maestro):
    colores_pestañas = {nombre: PatternFill("solid", fgColor=config["color"]) for nombre, config in constantes.PESTANAS.items()}
    color_encabezado = PatternFill("solid", fgColor="FF0000")
    fuente_encabezado = Font(bold=True, color="FFFFFF")
    negrita = Font(bold=True)

    ws = wb.create_sheet("RESUMEN")
    ws.append(["Bloque Original", "Resultado"])
    for celda in ws[1]:
        celda.font = negrita

    for red_original, contenido in reporte_maestro.items():
        ws.append([str(red_original), ""])
        for celda in ws[ws.max_row]:
            celda.fill = color_encabezado
            celda.font = fuente_encabezado

        bloques = contenido.get("bloques", {})

        # Agrupar por resultado dinámicamente
        por_resultado = {}
        for b, datos in bloques.items():
            resultado = datos["resultado"]
            por_resultado.setdefault(resultado, []).append(b)

        for resultado, lista in por_resultado.items():
            for bloque in juntar_sub_bloques(lista):
                ws.append([bloque, resultado])
                
                color = colores_pestañas.get(resultado, PatternFill("solid", fgColor="FFFFFF"))
                for celda in ws[ws.max_row]:
                    celda.fill = color

        ws.append([])
