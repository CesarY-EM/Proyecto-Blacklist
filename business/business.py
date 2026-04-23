import asyncio
import ipaddress
import ipaddress
import random
import threading


from concurrent.futures import ThreadPoolExecutor
from ipaddress import IPv4Network
from netaddr import cidr_merge
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pydnsbl import DNSBLIpChecker
from pydnsbl.providers import Provider, BASE_PROVIDERS
from constants import constantes
from models import models


_local= threading.local()

def juntar_bloques(lista_ips):
    return [str(res) for res in cidr_merge(lista_ips)]

def generar_excel_reporte(reporte_maestro):
    fill_red = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_font = Font(bold=True)

    nombre_archivo = f"reporte_prueba_v2.xlsx"
    
    wb = Workbook()

    negrita = Font(bold=True)
    rojo    = PatternFill("solid", fgColor="FFB3B3")
    verde   = PatternFill("solid", fgColor="B3FFB3")
    naranja = PatternFill("solid", fgColor="FFD9B3")

    todos_los_dominios = set()

    for datos in reporte_maestro.values():
        for bloque in datos.get("bloques", {}).values():

            if bloque.get("resultado") != "AUDITORIA":
                continue

            for ip in bloque.get("ips") or []:
                if not isinstance(ip, dict):
                    continue

                dom = ip.get("dominios", "")
                if dom:
                    todos_los_dominios.update(dom.split(", "))
    
    columnas_dnsbl = sorted(todos_los_dominios)


    # Pestaña RESUMEN
    ws_resumen = wb.active
    ws_resumen.title = "Resultados generales"
    ws_resumen.append(["Bloques", "Resultado"])

    for cell in ws_resumen[1]:
        cell.font = negrita


    for red_original, contenido in reporte_maestro.items():
        row = ws_resumen.append([str(red_original).strip(), ""])
        last_row = ws_resumen.max_row

        for col in range(1, 3):  # columnas A y B
            cell = ws_resumen.cell(row=last_row, column=col)
            cell.fill = fill_red
            cell.font = bold_font

        bloques = contenido.get("bloques", {}).items()

        bloques_bloqueo = [b for b, i in bloques if i.get("resultado") == "BLOQUEO"]
        bloques_limpio = [b for b, i in bloques if i.get("resultado") == "LIMPIO"]
        bloques_auditados = [b for b, i in bloques if i.get("resultado") == "AUDITORIA"]

        for bloque in juntar_bloques(bloques_bloqueo):
            ws_resumen.append([bloque, "BLOQUEO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = rojo

        for bloque in juntar_bloques(bloques_limpio):
            ws_resumen.append([bloque, "LIMPIO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = verde

        for bloque in juntar_bloques(bloques_auditados):
            ws_resumen.append([bloque, "AUDITORIA"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = naranja

        ws_resumen.append([])
        
    # Pestaña BLOQUEO
    ws_bloqueo = wb.create_sheet("BLOQUEO")
    ws_bloqueo.title = "BLOQUEO"
    ws_bloqueo.append(["Bloque", "Resultado"])
    for cell in ws_bloqueo[1]:
        cell.font = negrita

    
    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "BLOQUEO":
                continue
            ws_bloqueo.append([bloque, "BLOQUEO"])

    # Pestaña LIMPIO
    ws_limpio = wb.create_sheet("LIMPIO")
    ws_limpio.append(["Bloque", "Resultado"])

    for cell in ws_limpio[1]:
        cell.font = negrita

    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "LIMPIO":
                continue
            ws_limpio.append([bloque, "LIMPIO"])


    # Pestaña AUDITORIA
    ws_auditoria = wb.create_sheet("AUDITORIA")
    encabezados_auditoria = ["Bloque", "IP's", "resultado"] + columnas_dnsbl
    ws_auditoria.append(encabezados_auditoria)

    for cell in ws_auditoria[1]:
        cell.font = negrita

    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():

            if info.get("resultado") != "AUDITORIA":
                continue

            if not info.get("ips"):
                ws_auditoria.append(
                    [segmento, "N/A", "AUDITORIA"] + [""] * len(columnas_dnsbl) + [0]
                )
                continue

            for h in info.get("ips", []):
                if not h:
                    continue

                conteo = 0
                cols = []

                dominios = h.get("dominios", "")

                for dominio in columnas_dnsbl:
                    if dominio in dominios:
                        cols.append(1)
                        conteo += 1
                    else:
                        cols.append(0)

                fila = [segmento, h.get("ip"), "AUDITORIA"] + cols
                ws_auditoria.append(fila)

    wb.save(nombre_archivo)
    print("Reporte generado exitosamente")
    return nombre_archivo

def dividir_bloque(red):
    """
    Funcion que divide el bloque orginial en sub-bloques

    Args:
        string: Cadena que contiene el bloque original a dividir

    Returns:
        list: lista que contiene los subloques obtenidos
    """
    red = ipaddress.ip_network(red, strict = False)
    
    if red.prefixlen > 24:
        return list[red]
    if red.prefixlen >= 16:
        return list(red.subnets(new_prefix = constantes.PREFIJO_24))
    elif red.prefixlen < 16 and red.prefixlen >= 11:
        return list(red.subnets(new_prefix=constantes.PREFIJO_16))
    else: 
        return []
        # mensaje de bloque demasiado grande


def validar(bloques):

    """
    Funcion para validar si los bloques ingresados tienen el formato CIDR correcto

    Args:
        string: Bloques a validar

    Returns:
        list: Lista con los bloques que si tiene el formato correcto (se descartan los invalidos)
    """

    validos = []
    invalidos = []

    for bloque in bloques:
        try:
            red = IPv4Network(bloque.strip(), strict=False)
            validos.append(red)
        except ValueError as e:
            invalidos.append({"bloque": bloque.strip(), "error": str(e)})

    return validos


def obtener_checker():
    
    if not hasattr(_local, "checker"):
        dominios = BASE_PROVIDERS + constantes.PROVIDERS        
        _local.checker = DNSBLIpChecker(providers=dominios, timeout=2)
    return _local.checker

def verificar_ip(ip):

    """
    Funcion que verifica si direccion ingresada se encuentra en listas nehras

    Args:
        string: ip a verificar
        string: providers a agregar para contemplarlos para consulta

    Returns:
        dict: Contiene la direccion y dominios donde se encontro 
                (Si la direccion no se encontro en ningun dominio no se pasara nada)
    """

    try:
        checker = obtener_checker()
        print(f"Analizando: {ip}...", end="\r")
        resultado = checker.check(ip)

        if resultado.blacklisted:
            dominios = (
                ", ".join(resultado.detected_by.keys()) if resultado.detected_by else ""
            )
            return {"ip": ip, "dominios": dominios}
    except Exception as e:
        return None

def evaluar(resultados_muestra, total_muestra):
    
    """
    Funcion que evalua el resultado de las direcciones muestra

    Args:
        list: lista de las direcciones con hallazgos
        int: Numero del total de las direcciones muestra que se analizaron
        

    Returns:
        string: cadena que indica el resultado del porcentaje resultante
    """
    
    positivos = [r for r in resultados_muestra if r is not None]
    conteo_positivos = len(positivos)
    

    porcentaje = conteo_positivos / total_muestra
    
    if porcentaje >= constantes.UMBRAL:
        veredicto = "BLOQUEO"
        
    elif conteo_positivos == 0:
        veredicto = "LIMPIO"
    else:   
        veredicto = "AUDITORIA"
    
    return models.EvaluacionMuestra(
        resultado=veredicto,
        porcentaje=porcentaje,
        conteo=conteo_positivos
    )

async def consulta_exhaustiva(direcciones, loop, executor):
    """
    Funcion que evalua todas las direcciones de un bloque

    Args:
        list: lista donde contiene las direcciones de un bloque a analizar
        loop: administrador de tareas que organiza el tráfico de red
        executor: grupo de hilos de apoyo
        

    Returns:
        dict: diccionario que contiene la ip y los dominios donde se encontro
    """

    consulta_completa = [
        loop.run_in_executor(executor, verificar_ip, str(ip)) for ip in direcciones
    ]

    resultados = await asyncio.gather(*consulta_completa)
    return resultados
    

def obtener_muestra(red):
    todas = list(red.hosts())
    direcciones_muestra = []

    # Determinamos el tamaño de la muestra según el prefijo
    if red.prefixlen == 24:
        objetivo = constantes.MUESTRA_24
    elif red.prefixlen == 16:
        objetivo = constantes.MUESTRA_16
    else:
        objetivo = 5

    if objetivo > 0:
        if len(todas) <= objetivo:
            direcciones_muestra = todas[:]
        else:
            primera = todas[0]
            ultima = todas[-1]

            centro = todas[1:-1]
            
            centro_aleatorio = random.sample(centro, objetivo - 2)
            direcciones_muestra = [primera] + centro_aleatorio + [ultima]
    else:
        direcciones_muestra = []
    return direcciones_muestra

async def analizar_bloques(bloque, loop, executor):

    """
    Funcion que obtiene direcciones individuales de cada bloque y consulta muestreo

    Args:
        string: cadena que contiene el bloque de donde se obtendran las direcciones individuales

    Returns:
        string: resultado del metodo analizar_bloque
    """
    
    red = ipaddress.ip_network(bloque, strict=False)
            
    muestra = obtener_muestra(red) 

    verificar_muestra = [
        loop.run_in_executor(executor, verificar_ip, str(ip)) for ip in muestra
    ]
    resultados_muestra = await asyncio.gather(*verificar_muestra)
    evaluacion = evaluar(resultados_muestra, len(muestra))
    
    if evaluacion.resultado == "BLOQUEO":
        print(f"{bloque} terminado")
        
        return models.ResultadoBloque(
            bloque = str(bloque),
            resultado = "BLOQUEO"
    )

    elif evaluacion.resultado == "LIMPIO":
        print(f"{bloque} terminado")
        return models.ResultadoBloque(
            bloque = str(bloque),
            resultado = "LIMPIO"
        )

    else:
        # AUDITORIA
        todas = list(red.hosts())
        print(f"{bloque} terminado")
        hallazgos = await consulta_exhaustiva(todas, loop, executor)
        hallazgos_verdaderos = [h for h in hallazgos if h is not None]
        return models.ResultadoBloque(
            bloque = str(bloque),
            resultado = "AUDITORIA",
            hallazgos = hallazgos_verdaderos
        )


async def consultar(bloques):
    """
    Funcion que orquesta el analisis de los sub-bloques

    Args:
        lista: Lista con los sub-bloques a analizar

    Returns:
        dict: Diccionario con resultados del analisis de los sub-bloques
    """
    
    loop = asyncio.get_running_loop() 
    executor = ThreadPoolExecutor(max_workers = constantes.MAX_WORKERS)
    try:
        print(">>> Iniciando análisis asíncrono...")

        tareas = [analizar_bloques(b, loop, executor) for b in bloques]
        resultados = await asyncio.gather(*tareas)
        
        reporte = {}
                
        for datos in resultados:
            bloque = datos.bloque 
            
            reporte[bloque] = {
                "ips": datos.hallazgos,
                "resultado": datos.resultado
            }
    
        return reporte
    
    finally:
        print(">>> Finalizando hilos de trabajo...")
        try:
            await loop.run_in_executor(None, executor.shutdown, True)
        except (ValueError, RuntimeError):
            pass

async def comprobar_subredes_blacklist(subredes):

    """
    Funcion principal, es el "orquestador"

    Args:
        string: redes a analizar

    Returns:
        string: None
    """
    resultados = {}

    respuesta = False
    validos = validar(subredes)

    if not validos:
        print("error: Direcciones no validas")
        return

    

    for red in validos:
        bloques = [str(b) for b in dividir_bloque(red)]

        for bloque in bloques:
            print(bloque)

        try:
            respuesta = await consultar(bloques)
            
        except Exception as e:
            print(f"error:Error consultar -  {e}")
            return None
        
        if respuesta is None:
            print("error: consultar no devolvio resultados")
            return None
        
        resultados[str(red)] = {
        "bloques": respuesta
        }
        
    
    print("Generando reporte final")
    generar_excel_reporte(resultados)
        
    return respuesta