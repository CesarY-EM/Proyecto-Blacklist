import asyncio

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from business.business import validar
from business.business import consultar
from netaddr import IPNetwork, IPSet, cidr_merge


def juntar_bloques(lista_ips):
    return [str(res) for res in cidr_merge(lista_ips)]


def generar_excel_reporte(reporte_maestro):
    fill_red = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_font = Font(bold=True)

    nombre_archivo = f"reporte_prueba_v2.xlsx"
    
    wb = Workbook()

    negrita = Font(bold=True)
    rojo    = PatternFill("solid", fgColor="FFB3B3")
    verde   = PatternFill("solid", fgColor="B3FFB3")
    naranja = PatternFill("solid", fgColor="FFD9B3")

    todos_los_dominios = set()

    for datos in reporte_maestro.values():
        for bloque in datos.get("bloques", {}).values():

            if bloque.get("resultado") != "AUDITORIA":
                continue

            for ip in bloque.get("ips") or []:
                if not isinstance(ip, dict):
                    continue

                dom = ip.get("dominios", "")
                if dom:
                    todos_los_dominios.update(dom.split(", "))
    
    columnas_dnsbl = sorted(todos_los_dominios)


    # Pestaña RESUMEN
    ws_resumen = wb.active
    ws_resumen.title = "Resultados generales"
    ws_resumen.append(["Bloques", "Resultado"])

    for cell in ws_resumen[1]:
        cell.font = negrita


    for red_original, contenido in reporte_maestro.items():
        row = ws_resumen.append([str(red_original).strip(), ""])
        last_row = ws_resumen.max_row

        for col in range(1, 3):  # columnas A y B
            cell = ws_resumen.cell(row=last_row, column=col)
            cell.fill = fill_red
            cell.font = bold_font

        bloques = contenido.get("bloques", {}).items()

        bloques_bloqueo = [b for b, i in bloques if i.get("resultado") == "BLOQUEO"]
        bloques_limpio = [b for b, i in bloques if i.get("resultado") == "LIMPIO"]
        bloques_auditados = [b for b, i in bloques if i.get("resultado") == "AUDITORIA"]

        for bloque in juntar_bloques(bloques_bloqueo):
            ws_resumen.append([bloque, "BLOQUEO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = rojo

        for bloque in juntar_bloques(bloques_limpio):
            ws_resumen.append([bloque, "LIMPIO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = verde

        for bloque in juntar_bloques(bloques_auditados):
            ws_resumen.append([bloque, "AUDITORIA"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = naranja

        ws_resumen.append([])
        
    # Pestaña BLOQUEO
    ws_bloqueo = wb.create_sheet("BLOQUEO")
    ws_bloqueo.title = "BLOQUEO"
    ws_bloqueo.append(["Bloque", "Resultado"])
    for cell in ws_bloqueo[1]:
        cell.font = negrita

    
    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "BLOQUEO":
                continue
            ws_bloqueo.append([bloque, "BLOQUEO"])

    # Pestaña LIMPIO
    ws_limpio = wb.create_sheet("LIMPIO")
    ws_limpio.append(["Bloque", "Resultado"])

    for cell in ws_limpio[1]:
        cell.font = negrita

    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "LIMPIO":
                continue
            ws_limpio.append([bloque, "LIMPIO"])


    # Pestaña AUDITORIA
    ws_auditoria = wb.create_sheet("AUDITORIA")
    encabezados_auditoria = ["Bloque", "IP's", "resultado"] + columnas_dnsbl
    ws_auditoria.append(encabezados_auditoria)

    for cell in ws_auditoria[1]:
        cell.font = negrita

    for segmento, datos in reporte_maestro.items():
        for bloque, info in datos.get("bloques", {}).items():

            if info.get("resultado") != "AUDITORIA":
                continue

            if not info.get("ips"):
                ws_auditoria.append(
                    [segmento, "N/A", "AUDITORIA"] + [""] * len(columnas_dnsbl) + [0]
                )
                continue

            for h in info.get("ips", []):
                if not h:
                    continue

                conteo = 0
                cols = []

                dominios = h.get("dominios", "")

                for dominio in columnas_dnsbl:
                    if dominio in dominios:
                        cols.append(1)
                        conteo += 1
                    else:
                        cols.append(0)

                fila = [segmento, h.get("ip"), "AUDITORIA"] + cols
                ws_auditoria.append(fila)

    wb.save(nombre_archivo)
    print(f"Reporte generado exitosamente")
    return nombre_archivo

def dividir_bloque(red):
    """
    Funcion que divide el bloque orginial en sub-bloques

    Args:
        string: Cadena que contiene el bloque original a dividir

    Returns:
        list: lista que contiene los subloques obtenidos
    """
    
    if red.prefixlen > 24:
        return red
    if red.prefixlen >= 16:
        return list(red.subnets(new_prefix=24))
    elif red.prefixlen < 16 and red.prefixlen >= 11:
        return list(red.subnets(new_prefix=16))
    else: 
        return ""
        # mensaje de bloque demasiado grande


async def comprobar_subredes_blacklist(subredes):

    """
    Funcion principal, es el "orquestador"

    Args:
        string: redes a analizar

    Returns:
        string: None
    """
    resultados = {}

    respuesta = False
    validos = validar(subredes)

    if not validos:
        print("error: Direcciones no validas")
        return


    for red in subredes:
        print(type)
        bloques = dividir_bloque(red)

        for bloque in bloques:
            print(bloque)

        try:
            respuesta = await consultar(bloques)
            
        except Exception as e:
            print(f"error:Error consultar -  {e}")
            return None
        
        if respuesta is None:
            print("error: consultar no devolvio resultados")
            return None
        
        resultados[str(red)] = {
        "bloques": respuesta
        }
        
    
    print("Generando reporte final")
    generar_excel_reporte(resultados)
        
    return respuesta

if __name__ == "__main__":

    bloques = ["200.67.0.0/21"]

    respuesta = asyncio.run(comprobar_subredes_blacklist(bloques))
    
    if respuesta is None:
        print("No se pudo completar el análisis")  