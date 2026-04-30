import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from netaddr import cidr_merge


sys.path.append("/home/ngsop/lilaApp/plugins/utilidadesPlugins")
from loggingConfig import LoggerFileConfig
from constantesPlugins import LOG_CONFIG_FILES
logging = LoggerFileConfig().crearLogFile(LOG_CONFIG_FILES.get("blacklist_check"))

from constants import constantes

def juntar_sub_bloques(lista_ips):

    return [str(res) for res in cidr_merge(lista_ips)]


    fill_red = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    bold_font = Font(bold=True)

    nombre_archivo = "reporte_prueba_cl.xlsx"
    
    wb = Workbook()

    negrita = Font(bold=True)
    rojo    = PatternFill("solid", fgColor="FFB3B3")
    verde   = PatternFill("solid", fgColor="B3FFB3")
    naranja = PatternFill("solid", fgColor="FFD9B3")

    todos_los_dominios = set()

    for datos in reporte_general.values():
        for bloque in datos.get("bloques", {}).values():

            if bloque.get("resultado") != "AUDITORIA":
                continue

            for ip in bloque.get("ips") or []:
                if not isinstance(ip, dict):
                    continue

                dom = ip.get("dominios", "")
                if dom:
                    todos_los_dominios.update(dom.split(", "))
    
    columnas_dnsbl = sorted(todos_los_dominios)


    # Pestaña RESUMEN
    ws_resumen = wb.active
    ws_resumen.title = "Resultados generales"
    ws_resumen.append(["Bloques", "Resultado"])

    for cell in ws_resumen[1]:
        cell.font = negrita


    for red_original, contenido in reporte_general.items():
        row = ws_resumen.append([str(red_original).strip(), ""])
        last_row = ws_resumen.max_row

        for col in range(1, 3):  # columnas A y B
            cell = ws_resumen.cell(row=last_row, column=col)
            cell.fill = fill_red
            cell.font = bold_font

        bloques = contenido.get("bloques", {}).items()

        bloques_bloqueo = [b for b, i in bloques if i.get("resultado") == "BLOQUEO"]
        bloques_limpio = [b for b, i in bloques if i.get("resultado") == "LIMPIO"]
        bloques_auditados = [b for b, i in bloques if i.get("resultado") == "AUDITORIA"]

        for bloque in juntar_sub_bloques(bloques_bloqueo):
            ws_resumen.append([bloque, "BLOQUEO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = rojo

        for bloque in juntar_sub_bloques(bloques_limpio):
            ws_resumen.append([bloque, "LIMPIO"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = verde

        for bloque in juntar_sub_bloques(bloques_auditados):
            ws_resumen.append([bloque, "AUDITORIA"])
            for cell in ws_resumen[ws_resumen.max_row]:
                cell.fill = naranja

        ws_resumen.append([])
        
    # Pestaña BLOQUEO
    ws_bloqueo = wb.create_sheet("BLOQUEO")
    ws_bloqueo.title = "BLOQUEO"
    ws_bloqueo.append(["Bloque", "Resultado"])
    for cell in ws_bloqueo[1]:
        cell.font = negrita

    
    for segmento, datos in reporte_general.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "BLOQUEO":
                continue
            ws_bloqueo.append([bloque, "BLOQUEO"])

    # Pestaña LIMPIO
    ws_limpio = wb.create_sheet("LIMPIO")
    ws_limpio.append(["Bloque", "Resultado"])

    for cell in ws_limpio[1]:
        cell.font = negrita

    for segmento, datos in reporte_general.items():
        for bloque, info in datos.get("bloques", {}).items():
            if info.get("resultado") != "LIMPIO":
                continue
            ws_limpio.append([bloque, "LIMPIO"])


    # Pestaña AUDITORIA
    ws_auditoria = wb.create_sheet("AUDITORIA")
    encabezados_auditoria = ["Bloque", "IP's", "resultado"] + columnas_dnsbl
    ws_auditoria.append(encabezados_auditoria)

    for cell in ws_auditoria[1]:
        cell.font = negrita

    for segmento, datos in reporte_general.items():
        for bloque, info in datos.get("bloques", {}).items():

            if info.get("resultado") != "AUDITORIA":
                continue

            if not info.get("ips"):
                ws_auditoria.append(
                    [segmento, "N/A", "AUDITORIA"] + [""] * len(columnas_dnsbl) + [0]
                )
                continue

            for h in info.get("ips", []):
                if not h:
                    continue

                conteo = 0
                cols = []

                dominios = h.get("dominios", "")

                for dominio in columnas_dnsbl:
                    if dominio in dominios:
                        cols.append(1)
                        conteo += 1
                    else:
                        cols.append(0)

                fila = [segmento, h.get("ip"), "AUDITORIA"] + cols
                ws_auditoria.append(fila)

    wb.save(nombre_archivo)
    print("Reporte generado exitosamente")
    return nombre_archivo

def generar_reporte(reporte_general):
    #nombre_archivo = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    nombre_archivo = "prueba.xlsx"
    logging.info(f"Generando reporte")
    nombre_archivo = ("Prueba_refactorizacion.xlsx")
    wb = Workbook()
    negrita = Font(bold=True)

    # Recolectar dominios
    todos_los_dominios = set()
    for contenido in reporte_general.values():
        for datos in contenido.get("bloques", {}).values():
            if datos["resultado"] == "AUDITORIA":
                for h in datos.get("ips", []) or []:
                    if h and isinstance(h.get("dominios"), str):
                        todos_los_dominios.update(h["dominios"].split(", "))
    columnas_dnsbl = sorted(todos_los_dominios)

    # Agregar columnas
    constantes.PESTANAS["AUDITORIA"]["columnas"] = ["Bloque", "IP", "Resultado"] + columnas_dnsbl

    #Generar pestañas
    hojas = {}
    primera = True
    for nombre, config in constantes.PESTANAS.items():
        ws = wb.active if primera else wb.create_sheet(nombre)
        if primera:
            ws.title = nombre
            primera = False

        hojas[nombre] = ws
        color = PatternFill("solid", fgColor=config["color"])

        ws.append(config["columnas"])
        for celda in ws[1]:
            celda.font = negrita

        for red_original, contenido in reporte_general.items():
            for segmento, datos in contenido.get("bloques", {}).items():
                if datos["resultado"] != nombre:
                    continue

                if nombre == "AUDITORIA":
                    if not datos.get("ips"):
                        fila = [segmento, "N/A", nombre] + [""] * len(columnas_dnsbl)
                        ws.append(fila)
                        for celda in ws[ws.max_row]:
                            celda.fill = color
                        continue

                    for h in datos["ips"]:
                        if not h:
                            continue
                        ws.append([segmento, h["ip"], nombre])

                        for celda in ws[ws.max_row]:
                            celda.fill = color
                else:
                    ws.append([segmento, nombre])
                    for celda in ws[ws.max_row]:
                        celda.fill = color

    #Pestaña RESUMEN
    generar_resumen(wb, reporte_general)

    wb.save(nombre_archivo)
    print(f"✅ Reporte generado exitosamente: {nombre_archivo}")
    return nombre_archivo

def generar_resumen(wb, reporte_maestro):
    colores_pestañas = {nombre: PatternFill("solid", fgColor=config["color"]) for nombre, config in PESTANAS.items()}
    color_encabezado = PatternFill("solid", fgColor="FF0000")
    fuente_encabezado = Font(bold=True, color="FFFFFF")
    negrita = Font(bold=True)

    ws = wb.create_sheet("RESUMEN")
    ws.append(["Bloque Original", "Resultado"])
    for celda in ws[1]:
        celda.font = negrita

    for red_original, contenido in reporte_maestro.items():
        ws.append([str(red_original), ""])
        for celda in ws[ws.max_row]:
            celda.fill = color_encabezado
            celda.font = fuente_encabezado

        bloques = contenido.get("bloques", {})

        # Agrupar por resultado dinámicamente
        por_resultado = {}
        for b, datos in bloques.items():
            resultado = datos["resultado"]
            por_resultado.setdefault(resultado, []).append(b)

        for resultado, lista in por_resultado.items():
            for bloque in juntar_sub_bloques(lista):
                ws.append([bloque, resultado])
                
                color = colores_pestañas.get(resultado, PatternFill("solid", fgColor="FFFFFF"))
                for celda in ws[ws.max_row]:
                    celda.fill = color

        ws.append([])
